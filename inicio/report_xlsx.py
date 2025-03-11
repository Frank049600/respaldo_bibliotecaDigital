from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Side, Border
import os, re
from django.conf import settings
from django.http import HttpResponse

def insert_header_image(sheet):
    """
    Inserta una imagen de encabezado en la hoja del libro de Excel.
    """
    # Ruta de la imagen
    imagen_path = os.path.join(settings.BASE_DIR, 'inicio', 'static', 'img', 'header_image_uts.jpg')

    # Verifica si la imagen existe
    if not os.path.exists(imagen_path):
        raise FileNotFoundError(f"La imagen no existe en la ruta: {imagen_path}")

    # Inserta la imagen en el Excel
    img = Image(imagen_path)
    img.width = 700
    img.height = 100
    sheet.add_image(img, "B2")  # Inserta la imagen en la celda B2

def reporte_info(sheet, data):
    # Unión de celdas
    sheet.merge_cells('A11:H11')
    sheet['A11'].font = Font(color = 'FFFFFF', bold=True, size=12)
    sheet['A11'].fill = PatternFill('solid', start_color="d20606")
    sheet['A11'] = 'CONTROL DE REPORTES MENSUALES DE SERVICIOS: BIBLIOTECA.'

    sheet.merge_cells('A12:H12')
    sheet['A12'].font = Font(color = '000000', bold=True, size=12)
    sheet['A12'].fill = PatternFill('solid', start_color="d3c905")
    sheet['A12'] = 'CONSULTAS  EN EL CICLO DEL MES DE: ' + data['ciclo']

# def headers_by_tabla():
def get_borders(tipo):
    all_border = Border(
        left=Side(style="thin"),   # Borde izquierdo delgado
        top=Side(style="thin"),    # Borde superior delgado
        right=Side(style="thin"),   # Borde derecho delgado
        bottom=Side(style="thin")  # Borde inferior delgado
    )
    right_border = Border(
        top=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thin")
    )
    left_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        bottom=Side(style="thin")
    )
    top_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin")
    )
    bottom_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thin")
    )
    if tipo == 'all':
        regresa = all_border
    elif tipo == 'right':
        regresa = right_border
    elif tipo == 'left':
        regresa = left_border
    elif tipo == 'top':
        regresa = top_border
    elif tipo == 'bottom':
        regresa = bottom_border
    
    return regresa

# Crea la tabla de acervo
def table_acervo(sheet, data):
    """Función para la cración de registro para acervo

    Args:
        sheet (object): Instancia de Workbook
        data (array): Arreglo con la información repocilada
    """
    # configuraciones reusables reusables
    centrado = Alignment(horizontal='center', vertical='center')
    celdas = ['A', 'B', 'C', 'D', 'E', 'F']
    for c in celdas:
        sheet[c + '11'].fill = PatternFill('solid', start_color="0060df")
        sheet[c + '11'].font = Font(color = 'ffffff', bold=True, size=12)
        sheet[c + '12'].fill = PatternFill('solid', start_color="a0aab9")
        sheet[c + '13'].fill = PatternFill('solid', start_color="a0aab9")
    # Unión de celdas
    sheet.merge_cells('A11:D11')
    sheet.merge_cells('E11:F11')
    sheet['A11'] = 'REPORTE GENERAL DE ACERVO BIBLIOGRÁFICO:'
    sheet['E11'] = data['ciclo']

    # Crea encabezados de la tabla
    sheet['A12'].alignment = centrado
    sheet.merge_cells('A12:A13')
    sheet['A12'].border = get_borders('top')
    sheet['A13'].border = get_borders('bottom')
    sheet['A12'] = "No."
    
    # Agranda el tamaño de la celda A
    sheet.column_dimensions['A'].width = 10
    # Agranda el tamaño de la celda B
    sheet.column_dimensions['B'].width = 20
    sheet.merge_cells('B12:B13')
    sheet['B12'].border = get_borders('all')
    sheet['B13'].border = get_borders('all')
    sheet['B12'] = 'Área de conocimento'
    sheet['B12'].alignment = centrado
    # Se asignan las llaves de los grupo a las celdas
    cont_cell = 14
    cont_id = 1
    for key in data['conteo_ejemplares']:
        sheet[f"A{cont_cell}"] = cont_id
        sheet[f"A{cont_cell}"].alignment = centrado
        sheet[f"B{cont_cell}"] = key
        sheet[f"B{cont_cell}"].alignment = centrado
        cont_cell += 1
        cont_id += 1
    # Se asignan los valores en las celdas siguientes

    # ==> Se agregan titulos y volumenes de libros <==
    # Agranda el tamaño de la celda D
    sheet.column_dimensions['C'].width = 20
    # Unión de celdas
    sheet.merge_cells('C12:D12')
    # Se agregan bordes
    sheet['C12'].border = get_borders('left')
    sheet['D12'].border = get_borders('right')
    sheet['C12'] = 'Libros'
    sheet['C12'].alignment = centrado
    sheet['C13'].border = get_borders('all')
    sheet['C13'] = 'No.TITULOS'
    sheet['C13'].alignment = centrado

    cont_cell = 14
    totalizador_lib1 = 0
    for ord_libros in data['conteo_ejemplares']:
        if ord_libros in data['cantidad_libro']:
            sheet[f"C{cont_cell}"] = data['cantidad_libro'][ord_libros]
            totalizador_lib1 += data['cantidad_libro'][ord_libros]
        else:
            sheet[f"C{cont_cell}"] = 0
        sheet[f"C{cont_cell}"].alignment = centrado
        cont_cell += 1

    # Asigna número de volumenes
    sheet.column_dimensions['D'].width = 20
    sheet['D13'].border = get_borders('all')
    sheet['D13'] = 'No. DE VOLUMENES'
    sheet['D13'].alignment = centrado
    cont_cell = 14
    totalizador_lib2 = 0
    # for volms in data['volumenes_por_libro']:
    for ord_v_lib in data['conteo_ejemplares']:
        if ord_v_lib in data['volumenes_por_libro']:
            sheet[f"D{cont_cell}"] = data['volumenes_por_libro'][ord_v_lib]
            totalizador_lib2 += data['volumenes_por_libro'][ord_v_lib]
        else:
            sheet[f"D{cont_cell}"] = 0
        sheet[f"D{cont_cell}"].alignment = centrado
        cont_cell += 1

    # ==> Se agregan titulos y volumenes de "discos" <==]
    sheet.column_dimensions['E'].width = 20
    # Unión de celdas
    sheet.merge_cells('E12:F12')
    # Se agregan bordes
    sheet['E12'].border = get_borders('left')
    sheet['F12'].border = get_borders('right')
    sheet['E12'] = 'Discos'
    sheet['E12'].alignment = centrado
    sheet['E13'].border = get_borders('all')
    sheet['E13'] = 'No.TITULOS'
    sheet['E13'].alignment = centrado

    cont_cell = 14
    totalizador_disc1 = 0
    # for datas in data['cantidad_disco']:
    for ord_c_disc in data['conteo_ejemplares']:
        if ord_c_disc in data['cantidad_disco']:
            sheet[f"E{cont_cell}"] = data['cantidad_disco'][ord_c_disc]
            totalizador_disc1 += data['cantidad_disco'][ord_c_disc]
        else:
            sheet[f"E{cont_cell}"] = 0
        sheet[f"E{cont_cell}"].alignment = centrado
        cont_cell += 1
    # Asigna número de volumenes
    sheet.column_dimensions['F'].width = 20
    sheet['F13'].border = get_borders('all')
    sheet['F13'] = 'No. DE VOLUMENES'
    sheet['F13'].alignment = centrado

    cont_cell = 14
    totalizador_disc2 = 0
    # for volms in data['volumenes_por_disco']:
    for ord_v_disc in data['conteo_ejemplares']:
        if ord_v_disc in data['volumenes_por_disco']:
            sheet[f"F{cont_cell}"] = data['volumenes_por_disco'][ord_v_disc]
            totalizador_disc2 += data['volumenes_por_disco'][ord_v_disc]
        else:
            sheet[f"F{cont_cell}"] = 0
        sheet[f"F{cont_cell}"].alignment = centrado
        cont_cell += 1

    # ==> Tatalizadores
    for celda in celdas:
        # Color y estilo a la fila
        sheet[f"{celda}{cont_cell}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{cont_cell}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{cont_cell}"].border = get_borders('all')
        sheet[f"{celda}{cont_cell}"].alignment = centrado

    sheet[f"A{cont_cell}"] = 'Total'
    # Totalizador titulos libros
    sheet[f"C{cont_cell}"] = totalizador_lib1
    # Totalizador volumenes libros
    sheet[f"D{cont_cell}"] = totalizador_lib2
    # Totalizador titulos discos
    sheet[f"E{cont_cell}"] = totalizador_disc1
    # Totalizador volumenes discos
    sheet[f"F{cont_cell}"] = totalizador_disc2

def table_reporte_estadias(sheet, data):
    """Función para la creación de la tabla de reportes de estadías

    Args:
        sheet (object): Instancia de Workbook
        data (_type_): Arreglo con la información recopilada
    """
    # configuraciones reusables reusables
    centrado = Alignment(horizontal='center', vertical='center')
    # Agrega caracteristicas grupales
    celdas = ['A', 'B', 'C', 'D']
    for c in celdas:
        sheet[f"{c}11"].fill = PatternFill('solid', start_color="0060df")
        sheet[f"{c}11"].font = Font(color = 'ffffff', bold=True, size=12)
        sheet[f"{c}11"].border = get_borders('all')
        sheet[f"{c}12"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}12"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{c}12"].border = get_borders('all')
        sheet[f"{c}12"].alignment = centrado
    # Unión de celdas
    sheet.merge_cells('A11:C11')
    sheet['A11'] = 'REPORTE GENERAL DOCUEMTOS DE ESTADÍAS:'
    sheet['D11'] = data['ciclo']
    # Ajuste de ancho de celda
    sheet.column_dimensions['C'].width = 70
    # Crea encabezados de la tabla
    sheet.row_dimensions[12].height = 25
    sheet['A12'] = "No."
    # 
    sheet.merge_cells('B12:C12')
    sheet['B12'] = 'Área de conocimiento'
    # Se agregan numero de reportes por carrera
    sheet.column_dimensions['D'].width = 30
    sheet['D12'] = "No. Reportes"
    # Se agrega numéración de campos
    # Se agregan careras activas
    cont_cell = 13
    cont_id = 1
    totalizador1 = 0
    for carrera in data['conc_carreras']:
        sheet[f"A{cont_cell}"] = cont_id
        sheet[f"A{cont_cell}"].alignment = centrado
        # Se agregan campos de abreviaciones y nombres de las carreras
        sheet[f"B{cont_cell}"] = carrera
        sheet[f"B{cont_cell}"].alignment = centrado
        sheet[f"C{cont_cell}"] = data['conc_carreras'][carrera]
        sheet[f"C{cont_cell}"].alignment = centrado
        # Se agregan numero de reportes por carrera
        coincidencia = True
        for repo in data['estadias_reportes']:
            if re.search(re.escape(carrera), repo, re.IGNORECASE):
                sheet[f"D{cont_cell}"] = data['estadias_reportes'][repo]
                # Suma para totalizador
                totalizador1 += data['estadias_reportes'][repo]
                coincidencia = True
                break
            coincidencia = False
        # Si no existio ninguna coincidencia marca en 0
        if not coincidencia:
            sheet[f"D{cont_cell}"] = 0
        # Centrado de contenido
        sheet[f"D{cont_cell}"].alignment = centrado
        # Incremento de contadores                
        cont_cell += 1
        cont_id += 1

    # ==> Totalizadores
    for celda in celdas:
        sheet[f"{celda}{cont_cell}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{cont_cell}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{cont_cell}"].border = get_borders('all')
        sheet[f"{celda}{cont_cell}"].alignment = centrado

    sheet[f"A{cont_cell}"] = "Total"
    sheet[f"D{cont_cell}"] = totalizador1

    
def create_excel(data):
    """Crea un archivo Excel con una imagen de encabezado.

    Args:
        data (array): Arreglo con la información necesaria para la creación del archivo Excel

    Returns:
        object: Instancia de Workbook
    """
    # Crear el libro y la hoja
    book = Workbook()
    book.remove(book['Sheet']) # Se remueve la hoja que se activa por degfault

    acervo = book.create_sheet(title="Acervo")
    # Inserta la imagen en el encabezado
    insert_header_image(acervo)
    # Inserta información
    table_acervo(acervo, data)
    # Creación tabla de estadías 
    estadias = book.create_sheet(title="Estadías")
    insert_header_image(estadias)
    table_reporte_estadias(estadias, data)

    return book

def generate_report(data):
    try:
        # Crear el archivo Excel
        book = create_excel(data)
        # Configurar la respuesta HTTP para la descarga
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="Reporte mensual '+ data['ciclo'] +'.xlsx"'

        # Guardar el archivo Excel directamente en la respuesta
        book.save(response)

        return response
    except FileNotFoundError as e:
        # Manejo específico para errores relacionados con la imagen
        return HttpResponse(str(e), status=404)
    except Exception as e:
        # Manejo genérico de errores
        return HttpResponse(f"Error al generar el reporte: {str(e)}", status=500)
